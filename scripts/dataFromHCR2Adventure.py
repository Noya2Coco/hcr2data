# -*- coding: utf-8 -*-
from datetime import date
from PIL import Image, ImageDraw
import numpy as np
import pytesseract
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.drawing.image import Image as Img
import re
from fuzzywuzzy import fuzz, process
import dropbox
import os
from gestionNullCompteur import chargerCompteur

import makeGraph


"""
# Créez une instance de l'API Dropbox avec votre token d'accès
dbx = dropbox.Dropbox('sl.BnD39X1pfalNEM_MY8z1_YU2k0xhdiskHOWm4wbtpzXdbdGha_SlmGiFVwxsDiPSgum282IKxKahta-PvvxcrukLA7WvTtfNwoe0dIGrNuU23ZPtXgYIKnVtM-Dkj3rJQbsf995QZlAv')

# Chemin vers le dossier où vous souhaitez télécharger les images
dossier_destination = '/unchecked'

# Liste des fichiers dans votre compte Dropbox
fichiers = dbx.files_list_folder(dossier_destination).entries

# Téléchargez chaque fichier du dossier source vers un dossier de destination local
dossier_destination_local = '/var/www/DataFromHCR2Adventure/screenshots'


for fichier in fichiers:
    chemin_fichier_source = fichier.path_display
    nom_fichier = fichier.name

    # Vérifiez si le fichier est un fichier (et non un dossier ou autre)
    if isinstance(fichier, dropbox.files.FileMetadata):
        try:
            # Téléchargez le fichier dans le dossier de destination local
            dbx.files_download_to_file(f'{dossier_destination_local}/{nom_fichier}', chemin_fichier_source)
            print(f'Fichier téléchargé : {nom_fichier}')
        except Exception as e:
            print(f'Erreur lors du téléchargement de {nom_fichier} : {str(e)}')
    else:
        print(f'Ignoré : {nom_fichier} n\'est pas un fichier')
"""


# Créer une image du pseudo et une image du score des 8 joueurs du screen
def imageCrop():
    folderPath = "data/screenshots/fullScreenshots"
    files = os.listdir(folderPath)
    
    for fichier in files:
        filePath = os.path.join(folderPath, fichier)
        
        if os.path.isfile(filePath):
            image = Image.open(filePath)
            basename = os.path.basename(filePath)
            basename, ext = os.path.splitext(basename)
            
            coordsPseudoCrop = [900, 390, 1220, 470]
            coordsScoreCrop = [1740, 390, 1975, 470]
            
            for i in range(1, 9):
                pseudoCrop = image.crop(coordsPseudoCrop)
                scoreCrop = image.crop(coordsScoreCrop)
                
                pseudoCrop.save("data/screenshots/pseudoCropped/" + basename + "_(" + str(i) + ")" + ext)
                scoreCrop.save("data/screenshots/scoreCropped/" + basename + "_(" + str(i) + ")" + ext)
                
                coordsPseudoCrop[1] += 80
                coordsPseudoCrop[3] += 80
                coordsScoreCrop[1] += 80
                coordsScoreCrop[3] += 80 
    
    return


# Créer une version noir et blanc d'une image et l'enregistre dans dossier temporaire
# Return le path où elle a été sauvegardée
def turnImageBlackAndWhite(screenshotPath):
    dossier = os.path.basename(os.path.dirname(screenshotPath))
    basename = os.path.basename(screenshotPath)
    
    image = Image.open(screenshotPath)
    image = image.convert('L')
    image = image.point(blackOrWhite, '1')
    
    
    # Si c'est un pseudo
    if dossier == "pseudoCropped":
        tempImagePath = "data/screenshots/tempPseudoCropped/" + basename
        image.save(tempImagePath)
    
    # Si c'est un score
    elif dossier == "scoreCropped":
        tempImagePath = "data/screenshots/tempScoreCropped/" + basename
        image.save(tempImagePath)
        
    return tempImagePath


# Rendre pixel couleur >200 blanc et le reste noir
# Return 255 si >200, sinon return 0
def blackOrWhite(pixel):
    minColor = 200
    if pixel > minColor:
        return 255
    return 0


# Créer le style de la feuille excel
def makeSheetStyle(sheet):
    # Titres colonnes
    sheet["B1"] = "Dates"
    sheet["A2"] = "Players"

    # Couleurs fonds titres colonnes
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    sheet["A1"].fill = black_fill
    sheet["B1"].fill = yellow_fill
    sheet["A2"].fill = yellow_fill

    # Appliquer une bordure à la ligne 1
    border = Border(bottom=Side(style='thick'))
    for cell in sheet[1]:
        cell.border = border

    # Appliquer une bordure à la ligne 2
    border = Border(bottom=Side(style='double'))
    for cell in sheet[2]:
        cell.border = border

    # Appliquer des bordures à la colonne A
    border = Border(right=Side(style='thick'))
    for cell in sheet['A']:
        cell.border = border
        
    # Appliquer une bordure à la cellule A2
    border = Border(bottom=Side(style='thick'), right=Side(style='thick'))
    sheet['A2'].border = border

    return 


def find_best_match(extracted_name, list_real_names):
        # Utilise process.extractOne pour trouver la correspondance la plus proche
        best_match, score = process.extractOne(extracted_name, list_real_names)
        return best_match if score >= 50 else extracted_name


    
    
def playersData():
    
    # Charger l'image depuis le chemin
    #with open(screenshotPath, 'r', encoding='utf-8', errors='replace') as fichier:
    #    players_data = fichier.read()
    
    playersName = []
    playersNameScreen = []
    folderPath = "data/screenshots/pseudoCropped"
    files = os.listdir(folderPath)
    
    for fichier in files:
        filePath = os.path.join(folderPath, fichier)
        if os.path.isfile(filePath):
            image = Image.open(filePath)
            playerName = pytesseract.image_to_string(image)
            image.close()
            playerName = playerName.replace('\n', ' ').replace('\x0c', ' ').replace(' ', '')
            
            if playerName == '':
                playerName = "Null" + chargerCompteur()
                
            playersName.append(playerName)
            playersNameScreen.append(filePath)

    
    playersScore = []
    playersScoreScreen = []
    #folderPath = "/var/www/DataFromHCR2Adventure/data/screenshots/scoreCropped"
    # Apparement fonctionne mieux en noir et blanc #
    folderPath = "data/screenshots/tempScoreCropped"

    files = os.listdir(folderPath)
    
    for fichier in files:
        filePath = os.path.join(folderPath, fichier)
        if os.path.isfile(filePath):
            image = Image.open(filePath)
            playerScore = pytesseract.image_to_string(image)
            image.close()
            playerScore = playerScore.replace('\n', '').replace('\x0c', '').replace(' ', '')
            
            if playerScore == '':
                playerScore = "Null" + chargerCompteur()
            
            playersScore.append(playerScore)
            playersScoreScreen.append(filePath)
            

    playersData = dict(zip(playersName, playersScore))
    playersDataScreen = dict(zip(playersNameScreen, playersScoreScreen))

    #print(playersName)
    #print("\n" + str(playersScore))
    #print("\n" + str(playersData))
    
    #list_players_data = players_data.split("##########")
    list_real_names = [
        "Zorro",
        "|| Zeus",
        "EvgPlays",
        "DaBell",
        "RE|Brady",
        "trèkÿ",
        "Michi2",
        "Kacp",
        "Gamebro44",
        "|DC| 3DG3",
        "Papperellix",
        "Kosbow",
        "collier",
        #########################Kyank?
        "(PR)linus",
        #########################On break
        "Ykkönen",
        "Esko",
        "RS|KnightGP",
        "H|PA|Gabs",
        "Moo",
        "SP|Bill|RM",
        "Seby|PA|",
        "LUCAS",
        #Smoke99174 cheat
        "PR|Joplin~PA",
        "WaVe",
        #########################Rocket
        "PA|John",
        "Rrrr",
        "Jaqimo",
        "Mox",
        "@Aaditya.GG",
        #########################???
        "Champa",
        "Laser",
        "||Cobus",
        "Atlas",
        "Diesel",
        "Moi:D",
        "sebi",
        "Noya",
        "SingerYZL",
        "atef132",
        "I-Am-RoBo",
        "Can Tapanç",
        "RS|Galaxy",
        "PA|Ü|HyPêD",
        "Cube",
        "MaxiGaming",
        "|Eric|RR",
        #Top 51

        ]
       
    #players_data = []
    #for temp_players_data in list_players_data:
    #    # Modification du text reçu pour qu'il soit conforme
    #    cleaned_text = temp_players_data.replace("PRATIQUANT", "").replace("DE CANYONING", "")
    #    # Diviser le texte en lignes
    #    lines = cleaned_text.splitlines()
    #    non_empty_lines = [line.strip() for line in lines if line.strip()]

        # Diviser les lignes en deux parties (noms des joueurs et scores)
    #    midpoint = len(non_empty_lines) // 2
    #    player_names = non_empty_lines[:midpoint]
    #    player_scores = non_empty_lines[midpoint:]

 
        

    #    # Enlever les espaces possibles
    #    cleaned_scores = [score.replace(" ", "") for score in player_scores]
    #    cleaned_names = [find_best_match(name, list_real_names) for name in player_names]

    #    # Créer un dictionnaire à partir des noms des joueurs et des scores
    #    dict_players_data = dict(zip(cleaned_names, cleaned_scores))
    #    if dict_players_data:
    #        players_data.append(dict_players_data)
    
    #print(players_data)
    return playersData, playersDataScreen



def dataToSheet(dictPlayersData):
    excelPath = "data/logs/waitingData.xlsx"
    workbook = openpyxl.load_workbook(excelPath)
    sheet = workbook.active
    
    # Trouver la dernière colonne utilisée dans le tableau Excel
    last_column = sheet.max_column
    # Trouver la dernière lettre de la dernière colonne utilisée
    last_column_letter = get_column_letter(last_column)
    # Trouver l'index de la dernière lettre de la dernière colonne utilisée
    last_column_index = column_index_from_string(last_column_letter)

    # Calculer le nom de la nouvelle colonne en ajoutant 1 à la dernière colonne existante
    new_column_letter = get_column_letter(last_column_index +1)

    # Fusionner les cellules B1 à la nouvelle colonne
    sheet.merge_cells(f"B1:{new_column_letter}1")
    # Trouver la dernière ligne utilisée dans la colonne A (noms des joueurs)
    last_row = sheet.max_row
    # Créer une nouvelle colonne en ajoutant 1 à la dernière colonne existante
    sheet[f'{new_column_letter}2'] = date.today().strftime("%Y-%m-%d")

    # Parcourir les noms des joueurs et les scores
    for playerName, playerScore in dictPlayersData.items():
        # Parcourir la colonne A (noms des joueurs) pour trouver le nom correspondant
        player_found = False

        for cell in sheet['A']:
            if cell.value == playerName:
                # Insérer la valeur dans la nouvelle colonne à la même ligne
                sheet[f'{new_column_letter}{cell.row}'] = playerScore
                player_found = True
                break  # Sortir de la boucle une fois que le nom correspondant est trouvé

        if not player_found:
            last_row = sheet.max_row + 1
            sheet[f'A{last_row}'] = playerName
            sheet[f'{new_column_letter}{last_row}'] = playerScore

    makeSheetStyle(sheet)
    # Sauvegarder le classeur Excel
    workbook.save(excelPath)
    # Fermer le classeur Excel
    workbook.close()


def dataScreenToSheet(dictPlayersData):
    excelPath = "data/logs/waitingScreen.xlsx"
    workbook = openpyxl.load_workbook(excelPath)
    sheet = workbook.active
    
    # Trouver la dernière colonne utilisée dans le tableau Excel
    last_column = sheet.max_column
    # Trouver la dernière lettre de la dernière colonne utilisée
    last_column_letter = get_column_letter(last_column)
    # Trouver l'index de la dernière lettre de la dernière colonne utilisée
    last_column_index = column_index_from_string(last_column_letter)

    # Calculer le nom de la nouvelle colonne en ajoutant 1 à la dernière colonne existante
    new_column_letterP = get_column_letter(last_column_index +1)
    new_column_letterS = get_column_letter(last_column_index +2)
    
    
    
    
     
    # Fusionner les cellules B1 à la nouvelle colonne
    sheet.merge_cells(f"B1:{new_column_letterS}1")
    # Trouver la dernière ligne utilisée dans la colonne A (noms des joueurs)
    last_row = sheet.max_row
    # Créer une nouvelle colonne en ajoutant 1 à la dernière colonne existante
    sheet[f'{new_column_letterP}2'] = date.today().strftime("%Y-%m-%d")
    
    cell = 3
    # Parcourir les noms des joueurs et les scores
    for playerName, playerScore in dictPlayersData.items():
        img = Image.open(playerName)
        img.thumbnail((85, 20))
        
        img.save(playerName)
        
        img = Img(playerName)
        
        sheet.add_image(img, f'{new_column_letterP}{cell}')

        img = Image.open(playerScore)
        img.thumbnail((85, 20))
        img.save(playerScore)

        img = Img(playerScore)
        
        sheet.add_image(img, f'{new_column_letterS}{cell}')
        
        cell += 1


    makeSheetStyle(sheet)
    # Sauvegarder le classeur Excel
    workbook.save(excelPath)
    # Fermer le classeur Excel
    workbook.close()



imageCrop()
playersData, playersDataScreen = playersData()

# excel_filename = "/var/www/DataFromHCR2Adventure/data/logs/adventureData.xlsx"
excel_filename = "data/logs/waitingData.xlsx"

dataToSheet(playersData)
dataScreenToSheet(playersDataScreen)

makeGraph.createAllGraph()


###########################################
############ Tests et réponses ############
###########################################

#imagePath = turnImageBlackAndWhite("/var/www/DataFromHCR2Adventure/data/screenshots/pseudoCropped/screenentier_(8).jpg")
#imagePath = turnImageBlackAndWhite("/var/www/DataFromHCR2Adventure/data/screenshots/scoreCropped/screenentier_(8).jpg")


# Afficher le dictionnaire de données
#print("Player data :\n"+ str(playersData))
#print("Donnees extraites enregistrees")
